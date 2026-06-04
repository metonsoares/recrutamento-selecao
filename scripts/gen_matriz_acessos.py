from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

FONT = 'Arial'
wb = Workbook()
ws = wb.active
ws.title = 'Matriz de Acessos'

# Perfis (colunas a marcar). Master e Recrutador pré-preenchidos com o comportamento atual.
PROFILES = ['Master', 'Recrutador', 'RH', 'Gestor', 'Personalizado']
headers = ['Módulo', 'Submódulo / Tela', 'Ação', *PROFILES, 'Observação']

# Linhas: (modulo, submodulo, acao, master, recrutador, obs)
M = 'Sim'; N = 'Não'; B = ''  # branco para perfis a definir
rows = [
    ('Dashboard', 'Dashboard', 'Ver indicadores gerais', M, M, ''),

    ('Currículos', 'Candidatos (Kanban)', 'Ver lista / quadro', M, M, ''),
    ('Currículos', 'Candidatos (Kanban)', 'Mover card / mudar status', M, M, ''),
    ('Currículos', 'Candidatos (Kanban)', 'Editar vaga da candidatura', M, M, ''),
    ('Currículos', 'Candidatos (Kanban)', 'Buscar / filtrar', M, M, ''),
    ('Currículos', 'Ficha do candidato', 'Ver ficha (Resumo/Currículo)', M, M, ''),
    ('Currículos', 'Ficha - Ficha de Admissão', 'Ver / Editar', M, M, ''),
    ('Currículos', 'Ficha - Dados para contrato', 'Ver / Editar', M, M, ''),
    ('Currículos', 'Ficha - Documentos', 'Ver / Anexar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - Dados Bancários', 'Ver / Editar', M, M, ''),
    ('Currículos', 'Ficha - Férias', 'Ver / Lançar / Editar', M, M, ''),
    ('Currículos', 'Ficha - Advertências', 'Ver / Criar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - Atestados', 'Ver / Criar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - Contracheques', 'Ver / Anexar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - Folhas de ponto', 'Ver / Anexar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - ASOs', 'Ver / Editar', M, M, ''),
    ('Currículos', 'Ficha - Pesquisas de clima', 'Ver / Adicionar pesquisa', M, M, ''),
    ('Currículos', 'Ficha - Pesquisas de clima', 'Remover pesquisa', M, N, 'Hoje só Master'),
    ('Currículos', 'Ficha - Registros', 'Ver / Criar / Excluir', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Analisar IA', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Visualizar Teste Cultural', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Check Processos (background)', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Convidar para entrevista (WhatsApp)', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Exportar PDF', M, M, ''),
    ('Currículos', 'Ficha - Ações', 'Desligar funcionário', M, N, 'Hoje só Master'),
    ('Currículos', 'Ficha - Ações', 'Excluir candidato', M, N, 'Hoje só Master'),

    ('Currículos', 'Agenda de entrevistas', 'Ver agenda (por entrevistador / por dia)', M, M, ''),
    ('Currículos', 'Agenda de entrevistas', 'Configurar locais', M, M, ''),
    ('Currículos', 'Agenda de entrevistas', 'Configurar entrevistadores e janelas', M, M, ''),
    ('Currículos', 'Agenda de entrevistas', 'Remover agendamento', M, M, ''),

    ('Currículos', 'Config. de currículos', 'Seções e perguntas (formulário)', M, N, 'Hoje só Master'),
    ('Currículos', 'Config. de currículos', 'Vagas', M, N, 'Hoje só Master'),
    ('Currículos', 'Config. de currículos', 'Teste cultural', M, N, 'Hoje só Master'),

    ('Colaboradores', 'Em contrato', 'Ver lista', M, M, ''),
    ('Colaboradores', 'Contratados', 'Ver lista', M, M, ''),
    ('Colaboradores', 'Intermitentes', 'Ver lista', M, M, ''),
    ('Colaboradores', 'Freelancers', 'Ver lista', M, M, ''),
    ('Colaboradores', 'Desligados', 'Ver lista', M, M, ''),

    ('Pesquisas de clima', 'Cadastrar pesquisas', 'Ver / Criar / Editar / Excluir', M, M, ''),
    ('Pesquisas de clima', 'Cadastrar pesquisas', 'Importar .docx (IA) / Gerar link/QR', M, M, ''),
    ('Pesquisas de clima', 'Ver resultados', 'Ver indicadores / Análise IA / Exportar PDF', M, M, ''),
    ('Pesquisas de clima', 'Ver resultados', 'Remover resposta', M, N, 'Hoje só Master'),

    ('Documentos da empresa', 'Documentos da empresa', 'Ver / Adicionar / Editar / Excluir', M, N, 'Hoje só Master'),

    ('Mensagens WhatsApp', 'Mensagens WhatsApp', 'Ver conversas e histórico', M, N, 'Hoje só Master'),
    ('Mensagens WhatsApp', 'Mensagens WhatsApp', 'Excluir conversa', M, N, 'Hoje só Master'),

    ('Relatórios', 'Relatórios', 'Ver relatórios', M, N, 'Hoje só Master'),

    ('Configurações', 'WhatsApp / Z-API', 'Configurar credenciais / auto-resposta / webhook / logs', M, N, 'Hoje só Master'),
    ('Configurações', 'Configuração IA', 'Configurar chaves / modelo / prompt', M, N, 'Hoje só Master'),
    ('Configurações', 'Empresa - Cadastro de empresa', 'Ver / Editar (CNPJ, filiais, centros de custo)', M, N, 'Hoje só Master'),
    ('Configurações', 'Empresa - Cultura da empresa', 'Ver / Editar (identidade, missão, cultura, branding)', M, N, 'Hoje só Master'),
    ('Configurações', 'Usuários - Perfil de usuário', 'Ver / Editar', M, N, 'Hoje só Master'),
    ('Configurações', 'Usuários - Cadastro de usuários', 'Criar / Editar / Excluir usuários', M, N, 'Hoje só Master'),
    ('Configurações', 'Kanban - Colunas', 'Configurar ordem / visibilidade', M, N, 'Hoje só Master'),
]

# Estilos
header_fill = PatternFill('solid', start_color='1A5C38')
header_font = Font(name=FONT, bold=True, color='FFFFFF', size=11)
mod_fill = PatternFill('solid', start_color='E8F0EB')
mod_font = Font(name=FONT, bold=True, size=10)
cell_font = Font(name=FONT, size=10)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
yes_fill = PatternFill('solid', start_color='D9F2E3')
no_fill = PatternFill('solid', start_color='F7DADA')

# Cabeçalho
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = border

# Linhas
r = 2
prev_mod = None
for (mod, sub, acao, master, recr, obs) in rows:
    show_mod = mod if mod != prev_mod else ''
    prev_mod = mod
    vals = [show_mod, sub, acao, master, recr, B, B, B, obs]
    ws.append(vals)
    for c in range(1, len(vals) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.font = cell_font
        if c <= 3 or c == len(vals):
            cell.alignment = left
        else:
            cell.alignment = center
        if show_mod and c == 1:
            cell.font = mod_font; cell.fill = mod_fill
        # cores nos perfis pré-preenchidos (Master col4, Recrutador col5)
        if c in (4, 5):
            if cell.value == 'Sim': cell.fill = yes_fill
            elif cell.value == 'Não': cell.fill = no_fill
    r += 1

last = r - 1

# Validação Sim/Não nas colunas de perfis (D..H)
dv = DataValidation(type='list', formula1='"Sim,Não"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'D2:H{last}')

# Larguras
widths = {'A': 18, 'B': 30, 'C': 42, 'D': 11, 'E': 12, 'F': 10, 'G': 10, 'H': 14, 'I': 22}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{last}'
ws.row_dimensions[1].height = 28

# Aba de instruções
ws2 = wb.create_sheet('Instruções')
notes = [
    ['Matriz de Acessos e Permissões — Sistema de Recrutamento (Brownie do Ton)'],
    [''],
    ['Como usar:'],
    ['- Cada linha é um módulo/ação do sistema.'],
    ['- Marque "Sim" ou "Não" em cada coluna de perfil (use a lista suspensa).'],
    ['- Master e Recrutador já vêm preenchidos com o comportamento ATUAL do sistema.'],
    ['- RH, Gestor e Personalizado estão em branco para você definir.'],
    ['- A coluna "Observação" indica itens hoje restritos ao Master.'],
    [''],
    ['Perfis atuais no sistema: apenas Master (acesso total) e Recrutador.'],
    ['As demais colunas servem para planejar novos perfis antes de implementar.'],
]
for n in notes:
    ws2.append(n)
ws2['A1'].font = Font(name=FONT, bold=True, size=13)
ws2.column_dimensions['A'].width = 90
for row in ws2.iter_rows():
    for cell in row:
        if cell.row > 1:
            cell.font = Font(name=FONT, size=10)

out = r'G:\Meu Drive\Brownie do Ton\18_TI\APP - IA\Pesquisa clima organizacional\Matriz_Acessos_Permissoes.xlsx'
wb.save(out)
print('SAVED', out, 'rows', last)
